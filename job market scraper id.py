# -*- coding: utf-8 -*-
"""
Job Market Scraper Indonesia - Analyst Roles

Install:
    pip install pandas requests beautifulsoup4 duckduckgo-search openpyxl lxml

Run:
    python job_market_scraper_id_10.py

Output Excel sheets:
    Jobs, Skill Frequency, Source Summary, Scraping Log
"""

import os
import re
import time
import random
from datetime import datetime
from urllib.parse import urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from duckduckgo_search import DDGS
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ==================================================
# CONFIG
# ==================================================
SAVE_FOLDER = r"B:\01 GOLDINA RAINA\07 CODE\0001 OUTPUT DATA HERE"

KEYWORDS = [
    "Data Analyst Indonesia job",
    "Business Intelligence Analyst Indonesia job",
    "Reporting Analyst Indonesia job",
    "Data Scientist Indonesia job",
    "Data Analyst Jakarta job",
    "Data Analyst remote Indonesia job",
]

SITE_QUERIES = [
    "site:glints.com Data Analyst Indonesia",
    "site:kalibrr.com Data Analyst Indonesia",
    "site:jobstreet.co.id Data Analyst Indonesia",
    "site:id.indeed.com Data Analyst Indonesia",
    "site:linkedin.com/jobs Data Analyst Indonesia",
]

MAX_RESULTS_PER_QUERY = 25
REQUEST_TIMEOUT = 20
SLEEP_MIN = 1.2
SLEEP_MAX = 3.0

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/137.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
}

SKILL_KEYWORDS = [
    "SQL", "Python", "R", "Excel", "Google Sheets", "Power BI", "Tableau",
    "Looker", "Looker Studio", "Google Data Studio", "BigQuery", "PostgreSQL",
    "MySQL", "SQL Server", "ETL", "Data Warehouse", "Dashboard", "Reporting",
    "Data Visualization", "Statistics", "Machine Learning", "Pandas", "NumPy",
    "Spark", "Airflow", "dbt", "Google Analytics", "GA4", "A/B Testing",
    "Business Intelligence", "Data Modeling", "Data Cleaning", "Data Mining",
]

SECTION_KEYWORDS = [
    "requirement", "requirements", "qualification", "qualifications",
    "kualifikasi", "persyaratan", "minimum qualifications", "preferred qualifications",
    "what you will do", "responsibilities", "job description", "deskripsi pekerjaan",
    "tanggung jawab", "skills", "skill", "keahlian",
]

# ==================================================
# HELPERS
# ==================================================
def clean_text(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def normalize_url(url):
    return (url or "").split("#")[0].strip()


def get_domain(url):
    try:
        return urlparse(url).netloc.replace("www.", "")
    except Exception:
        return ""


def contains_job_signal(text):
    t = (text or "").lower()
    signals = [
        "job", "jobs", "career", "careers", "lowongan", "loker",
        "apply", "lamar", "full-time", "full time", "hybrid", "remote",
        "qualification", "requirement", "kualifikasi", "responsibilities",
    ]
    return any(s in t for s in signals)

# ==================================================
# SEARCH
# ==================================================
def search_urls():
    queries = KEYWORDS + SITE_QUERIES
    rows, logs, seen = [], [], set()

    with DDGS() as ddgs:
        for query in queries:
            print("=" * 70)
            print(f"Searching: {query}")
            try:
                result_items = list(ddgs.text(query, max_results=MAX_RESULTS_PER_QUERY, region="id-id"))
                logs.append({"Step": "Search", "Query": query, "Status": "OK", "Message": f"{len(result_items)} result(s)", "Timestamp": datetime.now()})
            except Exception as e:
                result_items = []
                logs.append({"Step": "Search", "Query": query, "Status": "ERROR", "Message": str(e), "Timestamp": datetime.now()})
                print(f"Search error: {e}")

            for item in result_items:
                url = normalize_url(item.get("href") or item.get("url") or "")
                if not url or url in seen:
                    continue
                seen.add(url)
                rows.append({
                    "Search Query": query,
                    "Search Title": clean_text(item.get("title", "")),
                    "Search Snippet": clean_text(item.get("body", "")),
                    "URL": url,
                    "Domain": get_domain(url),
                })
            time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))
    return rows, logs

# ==================================================
# PAGE FETCH
# ==================================================
def get_page_info(url):
    try:
        response = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        status_code = response.status_code
        content_type = response.headers.get("Content-Type", "")
        if status_code >= 400:
            return {"ok": False, "status_code": status_code, "title": "", "meta_desc": "", "h1": "", "h2": "", "text": "", "html_len": len(response.text or ""), "error": f"HTTP {status_code}", "content_type": content_type}

        soup = BeautifulSoup(response.text, "html.parser")
        for tag in soup(["script", "style", "noscript", "svg", "footer", "nav"]):
            tag.decompose()

        title = clean_text(soup.title.text if soup.title else "")
        meta_desc = ""
        meta = soup.find("meta", attrs={"name": "description"})
        if meta and meta.get("content"):
            meta_desc = clean_text(meta.get("content"))

        h1 = clean_text(" | ".join([h.get_text(" ", strip=True) for h in soup.find_all("h1")[:3]]))
        h2 = clean_text(" | ".join([h.get_text(" ", strip=True) for h in soup.find_all("h2")[:5]]))
        page_text = clean_text(soup.get_text(" ", strip=True))

        return {"ok": True, "status_code": status_code, "title": title, "meta_desc": meta_desc, "h1": h1, "h2": h2, "text": page_text, "html_len": len(response.text or ""), "error": "", "content_type": content_type}
    except Exception as e:
        return {"ok": False, "status_code": None, "title": "", "meta_desc": "", "h1": "", "h2": "", "text": "", "html_len": 0, "error": str(e), "content_type": ""}

# ==================================================
# EXTRACTORS
# ==================================================
def extract_position(title, h1, search_title):
    for c in [h1, title, search_title]:
        c = clean_text(c)
        if not c:
            continue
        c = re.sub(r"\s*[-|–]\s*(Glints|Kalibrr|JobStreet|Indeed|LinkedIn).*", "", c, flags=re.I)
        c = re.sub(r"\s*\|.*", "", c)
        if len(c) >= 5:
            return c[:180]
    return ""


def extract_company(text, title=""):
    combined = f"{title} {text}"
    patterns = [
        r"(?:Company|Perusahaan|Employer|Hiring Company)\s*[:\-]\s*([A-Za-z0-9&.,'()\-\s]{2,80})",
        r"(?:at|di)\s+(PT\.?\s+[A-Za-z0-9&.,'()\-\s]{2,80})",
        r"(PT\.?\s+[A-Za-z0-9&.,'()\-\s]{2,80})",
    ]
    for pattern in patterns:
        m = re.search(pattern, combined, flags=re.I)
        if m:
            company = clean_text(m.group(1))
            company = re.split(r"\s{2,}| Posted | Apply | Lowongan | Jobs ", company)[0]
            return company[:120]
    if " - " in title:
        parts = [clean_text(p) for p in title.split(" - ") if clean_text(p)]
        if len(parts) >= 2:
            return parts[-1][:120]
    return ""


def extract_location(text):
    locations = ["Jakarta", "Bandung", "Surabaya", "Tangerang", "Bekasi", "Depok", "Bogor", "Yogyakarta", "Semarang", "Medan", "Bali", "Denpasar", "Makassar", "Remote", "Hybrid", "Indonesia", "Batam", "Malang", "Solo", "Palembang", "Balikpapan", "Sidoarjo"]
    found = [loc for loc in locations if re.search(rf"\b{re.escape(loc)}\b", text, flags=re.I)]
    return ", ".join(dict.fromkeys(found))


def extract_post_date(text):
    patterns = [
        r"\b\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4}\b",
        r"\b\d{1,2}\s+(?:Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+\d{4}\b",
        r"\b\d{4}-\d{2}-\d{2}\b",
        r"\b\d{1,2}/\d{1,2}/\d{4}\b",
        r"\bposted\s+\d+\s+(?:days?|hours?|weeks?)\s+ago\b",
        r"\b\d+\s+(?:hari|jam|minggu)\s+(?:yang\s+)?lalu\b",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I)
        if m:
            return clean_text(m.group(0))
    return ""


def extract_salary(text):
    patterns = [r"Rp\s?[\d\.\,]+\s?(?:-|–|to|sampai)?\s?Rp?\s?[\d\.\,]*", r"IDR\s?[\d\.\,]+\s?(?:-|–|to)?\s?IDR?\s?[\d\.\,]*"]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I)
        if m:
            return clean_text(m.group(0))[:100]
    return ""


def extract_experience(text):
    patterns = [r"(\d+\+?)\s*(?:years|year|yrs|yr)\s*(?:of)?\s*experience", r"pengalaman\s*(?:minimal|min\.?|at least)?\s*(\d+\+?)\s*tahun", r"min(?:imum)?\s*(\d+\+?)\s*(?:tahun|years|year)"]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I)
        if m:
            return clean_text(m.group(0))
    return ""


def extract_skills(text):
    found = []
    for skill in SKILL_KEYWORDS:
        pattern = r"(?<![A-Za-z0-9])" + re.escape(skill) + r"(?![A-Za-z0-9])"
        if re.search(pattern, text, flags=re.I):
            found.append(skill)
    return found


def extract_requirement(text):
    lower = text.lower()
    starts = [lower.find(k.lower()) for k in SECTION_KEYWORDS if lower.find(k.lower()) != -1]
    chunk = text[min(starts):min(starts) + 2500] if starts else text[:1800]
    stop_words = ["privacy policy", "terms of use", "similar jobs", "lowongan lainnya", "apply now"]
    low_chunk = chunk.lower()
    stop_positions = [low_chunk.find(s) for s in stop_words if low_chunk.find(s) != -1]
    if stop_positions:
        chunk = chunk[:min(stop_positions)]
    return clean_text(chunk)[:2500]


def classify_industry(text):
    t = text.lower()
    rules = {
        "Finance / Banking": ["bank", "fintech", "financial", "finance", "insurance", "asuransi"],
        "E-commerce / Retail": ["e-commerce", "ecommerce", "retail", "marketplace"],
        "Technology": ["technology", "software", "saas", "startup", "tech"],
        "FMCG": ["fmcg", "consumer goods", "food", "beverage"],
        "Consulting": ["consulting", "consultant"],
        "Telecommunication": ["telecommunication", "telekomunikasi", "telco"],
        "Healthcare": ["healthcare", "medical", "hospital", "kesehatan"],
        "Education": ["education", "edtech", "university", "school"],
    }
    for industry, keys in rules.items():
        if any(k in t for k in keys):
            return industry
    return ""


def extract_job_record(search_row, page, no):
    full_text = clean_text(" ".join([search_row.get("Search Title", ""), search_row.get("Search Snippet", ""), page.get("title", ""), page.get("meta_desc", ""), page.get("h1", ""), page.get("h2", ""), page.get("text", "")]))
    skills = extract_skills(full_text)
    return {
        "No": no,
        "Position": extract_position(page.get("title", ""), page.get("h1", ""), search_row.get("Search Title", "")),
        "Company": extract_company(full_text, page.get("title", "")),
        "Location": extract_location(full_text),
        "Industry Guess": classify_industry(full_text),
        "Salary": extract_salary(full_text),
        "Experience": extract_experience(full_text),
        "Skills Found": ", ".join(skills),
        "SQL": "Yes" if "SQL" in skills else "",
        "Python": "Yes" if "Python" in skills else "",
        "Excel": "Yes" if "Excel" in skills else "",
        "Power BI": "Yes" if "Power BI" in skills else "",
        "Tableau": "Yes" if "Tableau" in skills else "",
        "Requirement": extract_requirement(full_text),
        "Date Posted": extract_post_date(full_text),
        "Source Domain": search_row.get("Domain", ""),
        "Search Query": search_row.get("Search Query", ""),
        "Search Snippet": search_row.get("Search Snippet", ""),
        "Link": search_row.get("URL", ""),
        "Page Status": page.get("status_code", ""),
        "Page Text Length": len(page.get("text", "")),
        "Scraping Date": datetime.now(),
    }

# ==================================================
# OUTPUT
# ==================================================
def build_skill_frequency(df_jobs):
    rows = []
    if df_jobs.empty or "Skills Found" not in df_jobs.columns:
        return pd.DataFrame(columns=["Skill", "Count"])
    for skill in SKILL_KEYWORDS:
        count = df_jobs["Skills Found"].fillna("").str.contains(rf"\b{re.escape(skill)}\b", case=False, regex=True).sum()
        if count > 0:
            rows.append({"Skill": skill, "Count": int(count)})
    return pd.DataFrame(rows).sort_values("Count", ascending=False) if rows else pd.DataFrame(columns=["Skill", "Count"])


def style_excel(path):
    wb = load_workbook(path)
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    dark_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = dark_fill if ws.title in ["Jobs", "Skill Frequency"] else header_fill
            cell.font = white_font if ws.title in ["Jobs", "Skill Frequency"] else Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if ws.cell(1, cell.column).value in ["Requirement", "Search Snippet"]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 28
    wb.save(path)


def main():
    os.makedirs(SAVE_FOLDER, exist_ok=True)
    print("Starting job scraping...")
    start_time = time.time()

    search_rows, logs = search_urls()
    print(f"Total unique URL from search: {len(search_rows)}")

    jobs = []
    no = 1

    for idx, row in enumerate(search_rows, start=1):
        url = row["URL"]
        print(f"[{idx}/{len(search_rows)}] Fetching: {url}")
        page = get_page_info(url)
        logs.append({"Step": "Fetch", "Query": row.get("Search Query", ""), "URL": url, "Status": "OK" if page.get("ok") else "ERROR", "Message": page.get("error", ""), "HTTP Status": page.get("status_code", ""), "Text Length": len(page.get("text", "")), "Timestamp": datetime.now()})

        signal_text = " ".join([row.get("Search Title", ""), row.get("Search Snippet", ""), page.get("title", ""), page.get("text", "")[:1000]])
        if page.get("ok") and contains_job_signal(signal_text):
            jobs.append(extract_job_record(row, page, no))
            no += 1
            print("  OK job captured")
        else:
            print("  Skipped / not readable job page")
        time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

    df_jobs = pd.DataFrame(jobs)
    if df_jobs.empty:
        df_jobs = pd.DataFrame(columns=["No", "Position", "Company", "Location", "Industry Guess", "Salary", "Experience", "Skills Found", "SQL", "Python", "Excel", "Power BI", "Tableau", "Requirement", "Date Posted", "Source Domain", "Search Query", "Search Snippet", "Link", "Page Status", "Page Text Length", "Scraping Date"])

    df_logs = pd.DataFrame(logs)
    df_source_summary = df_jobs.groupby("Source Domain", as_index=False).agg(Jobs=("Position", "count")).sort_values("Jobs", ascending=False) if not df_jobs.empty else pd.DataFrame(columns=["Source Domain", "Jobs"])
    df_skill = build_skill_frequency(df_jobs)

    output_file = os.path.join(SAVE_FOLDER, f"Job_Market_Data_Analyst_ID_{datetime.now():%Y%m%d_%H%M}.xlsx")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_jobs.to_excel(writer, sheet_name="Jobs", index=False)
        df_skill.to_excel(writer, sheet_name="Skill Frequency", index=False)
        df_source_summary.to_excel(writer, sheet_name="Source Summary", index=False)
        df_logs.to_excel(writer, sheet_name="Scraping Log", index=False)
    style_excel(output_file)

    print("\nSELESAI")
    print(f"Total jobs captured: {len(df_jobs)}")
    print(f"Runtime: {time.time() - start_time:.2f} seconds")
    print(output_file)


if __name__ == "__main__":
    main()
